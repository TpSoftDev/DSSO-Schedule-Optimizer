# This file is responsible for generating time grids for a new hires class schedule
# The programmatic equivalent to manually highlighting a student's grid on paper
# Generates a new timetable in the project directory

import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import time

# Add the external directory to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
external_directory = os.path.join(current_dir, "..")
sys.path.append(external_directory)

from api_calls.workday_api.workday_api import getStudentSchedule
from utils.helperFunctions import convert_to_time

# Get the Excel file path from command-line arguments
if len(sys.argv) < 2:
    print("Error: The path to the Excel file must be provided as an argument.")
    sys.exit(1)

excel_file_path = sys.argv[1]
print(f"Excel file path: {excel_file_path}")

# Check if the file exists
if not os.path.exists(excel_file_path):
    print("Error: The file Timetable.xlsx does not exist at the specified path.")
    sys.exit(1)

# Load the workbook and access the active sheet
wb = load_workbook(excel_file_path)
ws = wb.active

# Iterate through each day to fill in the grid
def populateGrid(data):
    for course in data:
        fillInDay(course, "U", 3)
        fillInDay(course, "M", 4)
        fillInDay(course, "T", 5)
        fillInDay(course, "W", 6)
        fillInDay(course, "R", 7)
        fillInDay(course, "F", 8)
        fillInDay(course, "S", 9)

# Iterates through each time slot in the "day" row of the grid and determines whether or not to highlight that slot
def fillInDay(course, day, rowNum):
    max_col = ws.max_column
    hour = 6
    for col_num_outer in range(2, max_col + 1, 12):  # hour loop:
        minute = 0
        for col_num_inner in range(col_num_outer, col_num_outer + 12):
            currentTime = time(hour, minute)
            if not isAvailable(course, day, currentTime):
                cell = ws.cell(row=rowNum, column=col_num_inner)
                fillColor = PatternFill(start_color="98FF98", end_color="98FF98", fill_type="solid")
                cell.fill = fillColor
            minute += 5
        hour += 1

# Observes one time slot on the Excel sheet and determines whether or not the course meets during that time
def isAvailable(course, day, currentTime):
    days = course["meetingDays"]
    startClass = convert_to_time(course["start"])
    endClass = convert_to_time(course["end"])
    for char in days:
        if char == day:
            if currentTime >= startClass and currentTime < endClass:
                return False
    return True

# Clears the grid
def clearGrid():
    min_row = 3
    min_col = 2
    max_row = 9
    for row in range(min_row, max_row + 1):
        for col in range(min_col, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            fillColor = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill = fillColor

# Testing the functions
clearGrid()
data1 = getStudentSchedule(0)

if data1:
    populateGrid(data1)
else:
    print("Error fetching data from Workday")

# Save the workbook to the specified path
wb.save(excel_file_path)
